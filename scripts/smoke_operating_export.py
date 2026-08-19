import argparse
import os
import shutil
import sys
import tempfile
from pathlib import Path


def main():
    parser = argparse.ArgumentParser(
        description="Smoke-test the complete operating workbook export"
    )
    parser.add_argument("database", type=Path)
    args = parser.parse_args()

    with tempfile.TemporaryDirectory(prefix="operating_export_") as temp_dir:
        temp_path = Path(temp_dir)
        test_database = temp_path / "supplier_data.db"
        export_path = temp_path / "operating.xlsx"
        shutil.copy2(args.database, test_database)
        os.environ["SUPPLY_CHAIN_DB_PATH"] = str(test_database)
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

        import ttkbootstrap as ttk
        from openpyxl import load_workbook

        import database
        from pages import ImportExportPage
        import pages.import_export_page as module

        database.init_db()
        root = ttk.Window(themename="flatly")
        root.withdraw()
        try:
            host = ttk.Frame(root)
            host.pack()
            page = ImportExportPage(host)
            module.filedialog.asksaveasfilename = lambda **_kwargs: str(
                export_path
            )
            module.messagebox.showinfo = lambda *_args, **_kwargs: None
            page.export_operating_workbook()
            assert export_path.exists()
            workbook = load_workbook(export_path, read_only=True)
            assert workbook.sheetnames == [
                "项目",
                "合同",
                "合同项目分配",
                "结算",
                "销项发票",
                "回款记录",
                "成本",
            ]
            for sheet in workbook.worksheets:
                assert sheet.max_row >= 1
                assert sheet.max_column >= 1
            workbook.close()
        finally:
            root.destroy()

    print("Operating workbook export smoke test passed")


if __name__ == "__main__":
    main()
