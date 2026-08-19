import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox, filedialog
from services import (
    contract_service,
    cost_service,
    finance_service,
    master_data_service,
    project_service,
    procurement_service,
)
from datetime import datetime
from openpyxl import Workbook, load_workbook


class ImportExportPage:
    def __init__(self, parent):
        self.parent = parent
        self.build_ui()

    def build_ui(self):
        header = ttk.Frame(self.parent)
        header.pack(fill=X, pady=(0, 16))
        ttk.Label(header, text="数据导入导出", style="PageTitle.TLabel").pack(anchor=W)
        ttk.Label(
            header, text="通过标准 Excel 模板批量维护主数据和采购记录",
            style="PageSub.TLabel",
        ).pack(anchor=W, pady=(4, 0))

        # 供应商导出导入
        frame1 = ttk.Labelframe(self.parent, text="供应商数据", bootstyle=PRIMARY)
        frame1.pack(fill=X, pady=(0, 10), padx=0, ipady=4)
        ttk.Button(frame1, text="导出供应商到 Excel", bootstyle=INFO, command=self.export_suppliers).pack(side=LEFT, padx=10, pady=10)
        ttk.Button(frame1, text="从 Excel 导入供应商", bootstyle=SUCCESS, command=self.import_suppliers).pack(side=LEFT, padx=10, pady=10)

        # 产品导出导入
        frame2 = ttk.Labelframe(self.parent, text="材料与供应商报价", bootstyle=PRIMARY)
        frame2.pack(fill=X, pady=(0, 10), padx=0, ipady=4)
        ttk.Button(frame2, text="导出产品到 Excel", bootstyle=INFO, command=self.export_products).pack(side=LEFT, padx=10, pady=10)
        ttk.Button(frame2, text="从 Excel 导入产品", bootstyle=SUCCESS, command=self.import_products).pack(side=LEFT, padx=10, pady=10)

        # 采购记录导出导入
        frame3 = ttk.Labelframe(self.parent, text="采购记录", bootstyle=PRIMARY)
        frame3.pack(fill=X, pady=(0, 10), padx=0, ipady=4)
        ttk.Button(frame3, text="导出采购记录到 Excel", bootstyle=INFO, command=self.export_purchases).pack(side=LEFT, padx=10, pady=10)
        ttk.Button(frame3, text="从 Excel 导入采购记录", bootstyle=SUCCESS, command=self.import_purchases).pack(side=LEFT, padx=10, pady=10)

        frame4 = ttk.Labelframe(
            self.parent, text="经营数据归档", bootstyle=PRIMARY
        )
        frame4.pack(fill=X, pady=(0, 10), padx=0, ipady=4)
        ttk.Button(
            frame4,
            text="导出项目经营全量工作簿",
            bootstyle=INFO,
            command=self.export_operating_workbook,
        ).pack(side=LEFT, padx=10, pady=10)
        ttk.Label(
            frame4,
            text="项目、合同分配、结算、开票、回款和成本",
            style="PageSub.TLabel",
        ).pack(side=LEFT, padx=8)

        # 说明
        info = ttk.Label(self.parent, text="说明：采购数据已使用新版统一格式，正式采购需有供应商和产品，零星采购可只填商户与材料。项目名称不存在时，导入会自动建立项目。", wraplength=800, justify=LEFT)
        info.configure(style="PageSub.TLabel")
        info.pack(anchor=W, pady=(8, 0))

    def export_operating_workbook(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"项目经营全量_{datetime.now():%Y%m%d}.xlsx",
        )
        if not path:
            return
        workbook = Workbook()
        workbook.remove(workbook.active)

        def add_sheet(name, headers, rows):
            sheet = workbook.create_sheet(name)
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            sheet.freeze_panes = "A2"

        projects = project_service.list_projects()
        add_sheet(
            "项目",
            ["项目编码", "项目名称", "客户", "状态", "负责人", "计划开始", "计划结束"],
            [
                [
                    row["project_code"], row["name"], row["customer_name"],
                    row["status"], row["manager"], row["planned_start_date"],
                    row["planned_end_date"],
                ]
                for row in projects
            ],
        )
        add_sheet(
            "合同",
            ["合同编号", "合同名称", "客户", "类型", "签订日期", "合同金额", "已分配", "状态"],
            [
                [
                    row["contract_no"], row["name"], row["customer_name"],
                    contract_service.CONTRACT_TYPES[row["contract_type"]],
                    row["sign_date"], row["tax_inclusive_amount_minor"] / 100,
                    row["allocated_minor"] / 100,
                    contract_service.CONTRACT_STATUSES[row["status"]],
                ]
                for row in contract_service.list_contracts(include_void=True)
            ],
        )
        add_sheet(
            "合同项目分配",
            ["合同编号", "合同名称", "项目编码", "项目名称", "分配金额", "说明"],
            [
                [
                    row["contract_no"], row["contract_name"],
                    row["project_code"], row["project_name"],
                    row["allocated_amount_minor"] / 100, row["notes"],
                ]
                for row in contract_service.list_allocations()
            ],
        )
        add_sheet(
            "结算",
            [
                "结算编号", "日期", "项目", "合同", "结算金额",
                "已开票金额", "开票比例", "待开票金额", "依据",
            ],
            [
                [
                    row["settlement_no"], row["settlement_date"],
                    row["project_name"], row["contract_no"],
                    row["amount_minor"] / 100,
                    row["invoiced_minor"] / 100,
                    row["invoice_rate_percent"],
                    row["uninvoiced_minor"] / 100,
                    row["basis"],
                ]
                for row in contract_service.list_settlements()
            ],
        )
        add_sheet(
            "销项发票",
            [
                "发票号码", "日期", "项目", "合同", "收入确认",
                "购买方", "税率", "金额",
            ],
            [
                [
                    row["invoice_no"], row["invoice_date"], row["project_name"],
                    row["contract_no"], row["settlement_no"],
                    row["buyer_name_snapshot"],
                    row["tax_rate_bps"] / 100, row["amount_minor"] / 100,
                ]
                for row in finance_service.list_invoices()
            ],
        )
        add_sheet(
            "回款记录",
            ["回款单号", "日期", "项目", "合同", "付款方", "方式", "金额"],
            [
                [
                    row["receipt_no"], row["receipt_date"], row["project_name"],
                    row["contract_no"], row["payer_name_snapshot"],
                    row["payment_method"], row["allocated_amount_minor"] / 100,
                ]
                for row in finance_service.list_receipts()
            ],
        )
        project_names = {row["id"]: row["name"] for row in projects}
        add_sheet(
            "成本",
            ["日期", "项目", "来源单号", "分类", "往来单位或人员", "车辆/车牌", "来源", "金额"],
            [
                [
                    row["business_date"], row.get("allocation_project_names")
                    or project_names.get(row["project_id"], "待归集"),
                    row["source_no"], row["category"], row["counterparty"],
                    row.get("vehicle_no", ""), row["source_type"],
                    row["amount_minor"] / 100,
                ]
                for row in cost_service.list_cost_ledger()
            ],
        )
        workbook.save(path)
        messagebox.showinfo("导出成功", f"经营数据已归档到：\n{path}")

    def export_suppliers(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商"
        headers = [
            "ID", "工厂名称", "产品范围", "联系人", "价格水平", "交期",
            "质量情况", "是否做出口", "备注", "默认税率（%）"
        ]
        ws.append(headers)
        for row in master_data_service.list_suppliers(active_only=False):
            ws.append([row["id"], row["name"], row["category"], row["contact"], row["price_level"],
                       row["delivery"], row["quality"], row["export"], row["notes"],
                       row["default_tax_rate_percent"]])
        wb.save(path)
        messagebox.showinfo("成功", f"供应商数据已导出到：\n{path}")

    def import_suppliers(self):
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            data = {
                "name": row[1],
                "category": row[2] if row[2] else "钢材",
                "contact": row[3] if row[3] else "",
                "price_level": row[4] if row[4] else "中",
                "delivery": row[5] if row[5] else "一般",
                "quality": row[6] if row[6] else "良",
                "export": row[7] if row[7] else "否",
                "notes": row[8] if len(row) > 8 and row[8] else "",
                "default_tax_rate_percent": (
                    row[9] if len(row) > 9 and row[9] is not None else 0
                ),
            }
            master_data_service.create_supplier(data)
            count += 1
        messagebox.showinfo("成功", f"成功导入 {count} 条供应商数据")

    def export_products(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "产品"
        headers = [
            "ID", "供应商ID", "供应商名称", "产品名称", "规格", "单位",
            "材料单价（未税）", "税率（%）", "含税单价", "备注"
        ]
        ws.append(headers)
        for row in master_data_service.list_supplier_offers(active_only=False):
            ws.append([row["id"], row["supplier_id"], row["supplier_name"], row["name"],
                       row["specification"], row["unit"], row["price"],
                       row["tax_rate_percent"], row["tax_inclusive_price"], row["notes"]])
        wb.save(path)
        messagebox.showinfo("成功", f"产品数据已导出到：\n{path}")

    def import_products(self):
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        new_format = "税率（%）" in headers
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[3]:
                continue
            try:
                supplier_id = int(row[1]) if row[1] else None
                price = float(row[6]) if row[6] else 0
                tax_rate = (
                    float(row[7]) if new_format and row[7] is not None else None
                )
            except ValueError:
                continue
            if not supplier_id:
                continue
            data = {
                "supplier_id": supplier_id,
                "name": row[3],
                "specification": row[4] if row[4] else "",
                "unit": row[5] if row[5] else "",
                "price": price,
                "notes": (
                    row[9] if new_format and len(row) > 9 and row[9]
                    else row[7] if not new_format and len(row) > 7 and row[7]
                    else ""
                ),
            }
            if tax_rate is not None:
                data["tax_rate_percent"] = tax_rate
            master_data_service.create_supplier_offer(data)
            count += 1
        messagebox.showinfo("成功", f"成功导入 {count} 条产品数据")

    def export_purchases(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "采购记录"
        headers = [
            "ID", "采购单号", "采购类型", "采购日期", "项目", "供应商ID", "产品ID",
            "供应商/商户", "材料名称", "规格", "单位", "数量",
            "材料单价（未税）", "税率（%）", "含税单价", "未税材料额",
            "税额", "含税材料额", "运费", "计入项目成本",
            "成本类别", "支付方式", "支付状态", "票据状态", "经办人", "用途", "备注"
        ]
        ws.append(headers)
        for row in procurement_service.list_purchase_orders():
            ws.append([
                row["id"], row["order_no"], row["purchase_type"], row["purchase_date"],
                row["project_name"] or "", row["supplier_id"], row["product_id"],
                row["merchant_name_snapshot"], row["material_name_snapshot"],
                row["specification_snapshot"], row["unit_snapshot"], row["quantity"],
                row["material_unit_price_cents"] / 100, row["tax_rate_bps"] / 100,
                row["tax_inclusive_unit_price_cents"] / 100,
                row["material_amount_cents"] / 100, row["tax_amount_cents"] / 100,
                row["line_amount_cents"] / 100, row["freight_amount_cents"] / 100,
                row["project_cost_cents"] / 100, row["cost_category"],
                row["payment_method"], row["payment_status"],
                row["invoice_status"], row["purchaser"], row["purpose"],
                row["notes"] or row["item_notes"]
            ])
        wb.save(path)
        messagebox.showinfo("成功", f"采购记录已导出到：\n{path}")

    def import_purchases(self):
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        is_v2 = "采购类型" in headers and "供应商/商户" in headers
        has_tax_freight = "材料单价（未税）" in headers and "运费" in headers
        column = {name: index for index, name in enumerate(headers)}
        projects = {p["name"]: p["id"] for p in project_service.list_projects()}
        count = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if is_v2:
                    if not row or not row[2] or not row[7] or not row[8]:
                        skipped += 1
                        continue
                    project_name = str(row[4]).strip() if row[4] else ""
                    if project_name and project_name not in projects:
                        projects[project_name] = project_service.create_project({"name": project_name, "notes": "由采购 Excel 导入创建"})
                    purchase_type = row[2] if row[2] in ("正式采购", "零星采购") else "零星采购"
                    supplier_id = int(row[5]) if row[5] else None
                    product_id = int(row[6]) if row[6] else None
                    quantity = float(row[11] or 1)
                    unit_price = float(row[12] or 0)
                    amount = float(row[13] or quantity * unit_price)
                    if has_tax_freight:
                        tax_rate = float(row[column["税率（%）"]] or 0)
                        freight = float(row[column["运费"]] or 0)
                        cost_category = row[column["成本类别"]] or "材料费"
                        payment_method = row[column["支付方式"]] or "未记录"
                        payment_status = row[column["支付状态"]] or "未确认"
                        invoice_status = row[column["票据状态"]] or "未确认"
                        purchaser = row[column["经办人"]] or ""
                        purpose = row[column["用途"]] or ""
                        notes = row[column["备注"]] or ""
                    else:
                        tax_rate = None
                        freight = 0
                        cost_category = row[14] or "材料费"
                        payment_method = row[15] or "未记录"
                        payment_status = row[16] or "未确认"
                        invoice_status = row[17] or "未确认"
                        purchaser = row[18] or ""
                        purpose = row[19] or ""
                        notes = row[20] or ""
                    procurement_service.add_purchase_order({
                        "order_no": str(row[1]).strip() if row[1] else None,
                        "purchase_type": purchase_type,
                        "project_id": projects.get(project_name),
                        "supplier_id": supplier_id,
                        "merchant_name_snapshot": str(row[7]).strip(),
                        "purchase_date": str(row[3])[:10],
                        "payment_method": payment_method,
                        "payment_status": payment_status,
                        "invoice_status": invoice_status,
                        "purchaser": purchaser,
                        "freight_amount_cents": round(freight * 100),
                        "notes": notes,
                    }, {
                        "product_id": product_id, "material_name_snapshot": str(row[8]).strip(),
                        "specification_snapshot": row[9] or "", "unit_snapshot": row[10] or "",
                        "quantity": quantity,
                        "cost_category": cost_category,
                        "purpose": purpose, "notes": notes,
                        **(
                            {
                                "material_unit_price_cents": round(unit_price * 100),
                                "tax_rate_bps": round(tax_rate * 100),
                            }
                            if tax_rate is not None
                            else {
                                "unit_price_cents": round(unit_price * 100),
                                "line_amount_cents": round(amount * 100),
                            }
                        ),
                    })
                else:
                    # 兼容旧版 12 列模板，导入后直接进入新版正式采购。
                    if not row or not row[2] or not row[3]:
                        skipped += 1
                        continue
                    supplier_id, product_id = int(row[2]), int(row[3])
                    supplier = master_data_service.get_supplier_by_legacy_id(supplier_id)
                    product = master_data_service.get_supplier_offer_by_legacy_id(product_id)
                    if not supplier or not product:
                        skipped += 1
                        continue
                    project_name = str(row[10]).strip() if len(row) > 10 and row[10] else ""
                    if project_name and project_name not in projects:
                        projects[project_name] = project_service.create_project({"name": project_name, "notes": "由旧版采购 Excel 导入创建"})
                    quantity, unit_price = float(row[7] or 1), float(row[8] or 0)
                    amount = float(row[9] or quantity * unit_price)
                    procurement_service.add_purchase_order({
                        "purchase_type": "正式采购", "project_id": projects.get(project_name),
                        "supplier_id": supplier["id"], "merchant_name_snapshot": supplier["name"],
                        "purchase_date": str(row[1])[:10], "notes": row[11] if len(row) > 11 and row[11] else "",
                    }, {
                        "product_id": product["id"], "material_name_snapshot": product["name"],
                        "specification_snapshot": product["specification"], "unit_snapshot": product["unit"],
                        "quantity": quantity, "unit_price_cents": round(unit_price * 100),
                        "line_amount_cents": round(amount * 100), "cost_category": "材料费",
                    })
                count += 1
            except Exception:
                skipped += 1
        messagebox.showinfo("导入完成", f"成功导入 {count} 条，跳过 {skipped} 条。")


