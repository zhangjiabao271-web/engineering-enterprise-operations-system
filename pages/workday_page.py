import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox, filedialog, simpledialog, Listbox, EXTENDED, END as TK_END
from services import labor_service as db
from services import project_service
from datetime import date, datetime
from openpyxl import Workbook
from ui.components import DataTable, DatePicker, FilterBar, KpiCard, PageHeader
from ui.dialogs import add_form_actions, build_form_dialog, safe_init_loaders
from ui.theme import COLORS, SPACING

UNASSIGNED_LABEL = "待归集（不归入任何项目）"


class WorkdayDashboardPage:
    def __init__(self, parent):
        self.parent = parent
        self.month_var = ttk.StringVar(value=datetime.now().strftime("%Y-%m"))
        self.search_var = ttk.StringVar()
        self.overtime_hint_var = ttk.StringVar(value="其中加班 0 条 / 0 工天")
        self.kpi_vars = {
            "days": ttk.StringVar(value="0"),
            "amount": ttk.StringVar(value="¥0"),
            "workers": ttk.StringVar(value="0"),
            "sites": ttk.StringVar(value="0"),
        }
        self.build_ui()
        safe_init_loaders("人工与工天", [self.refresh_months, self.refresh_all])

    def build_ui(self):
        # ---- Header（组件化；月份选择跟随操作区）----
        month_box = ttk.Frame(self.parent)
        ttk.Label(month_box, text="月份", style="PageSub.TLabel").pack(side=LEFT, padx=(0, 8))
        self.month_combo = ttk.Combobox(
            month_box, textvariable=self.month_var, width=10, state="readonly"
        )
        self.month_combo.pack(side=LEFT)
        self.month_combo.bind("<<ComboboxSelected>>", lambda event: self.refresh_all())
        PageHeader(
            self.parent,
            "人工与工天",
            "按月查看工人出勤、工地去向和人工成本",
            actions=[
                month_box,
                ttk.Button(
                    self.parent, text="新增工天", bootstyle=SUCCESS,
                    command=self.open_log_dialog,
                ),
                ttk.Button(
                    self.parent, text="导出本月", bootstyle=INFO,
                    command=self.export_month,
                ),
                ttk.Button(
                    self.parent, text="新增工人", bootstyle=OUTLINE,
                    command=self.open_worker_dialog,
                ),
            ],
        )

        # ---- KPI 卡片行（4列，组件化，去掉角标图标）----
        self.static_hint_vars = {
            "amount": ttk.StringVar(value="较上月 --"),
            "workers": ttk.StringVar(value="人"),
            "sites": ttk.StringVar(value="个"),
        }
        kpi_specs = [
            ("本月总工天", self.kpi_vars["days"], self.overtime_hint_var),
            ("预计人工成本", self.kpi_vars["amount"], self.static_hint_vars["amount"]),
            ("出勤工人", self.kpi_vars["workers"], self.static_hint_vars["workers"]),
            ("施工工地", self.kpi_vars["sites"], self.static_hint_vars["sites"]),
        ]
        kpi_frame = ttk.Frame(self.parent)
        kpi_frame.pack(fill=X, pady=(0, SPACING["md"]))
        for index, (label, value_var, hint_var) in enumerate(kpi_specs):
            KpiCard(kpi_frame, label, value_var, hint_var).grid(
                row=0, column=index, sticky=EW,
                padx=(0 if index == 0 else 6, 0 if index == 3 else 6),
            )
            kpi_frame.columnconfigure(index, weight=1)

        # ---- 两栏 Overview ----
        overview = ttk.Panedwindow(self.parent, orient=HORIZONTAL)
        overview.pack(fill=X, pady=(0, SPACING["md"]))

        worker_card = ttk.Frame(overview, style="Card.TFrame", padding=14)
        site_card = ttk.Frame(overview, style="Card.TFrame", padding=14)
        overview.add(worker_card, weight=1)
        overview.add(site_card, weight=1)

        # 工人工天排行（组件化；排名与 top3 高亮在 refresh_ranks 的 mapper 中生成）
        ttk.Label(
            worker_card, text="工人工天排行", style="CardTitle.TLabel"
        ).pack(anchor=W, pady=(0, 8))
        self.worker_rank = DataTable(
            worker_card,
            specs=(
                ("rank", "排名", 40, CENTER),
                ("worker", "工人", 90, CENTER),
                ("trade", "工种", 90, CENTER),
                ("days", "工天", 65, CENTER),
                ("overtime", "加班工天", 70, CENTER),
                ("amount", "人工费", 85, CENTER),
            ),
            empty_text="本月暂无工人工天记录",
            stretch=("worker",),
        )

        # 工地投入分布
        ttk.Label(
            site_card, text="工地投入分布", style="CardTitle.TLabel"
        ).pack(anchor=W, pady=(0, 8))
        self.site_rank = DataTable(
            site_card,
            specs=(
                ("rank", "排名", 40, CENTER),
                ("site", "工地", 140, CENTER),
                ("workers", "人数", 55, CENTER),
                ("days", "工天", 65, CENTER),
                ("overtime", "加班工天", 70, CENTER),
                ("amount", "人工费", 85, CENTER),
            ),
            empty_text="本月暂无工地投入记录",
            stretch=("site",),
        )
        for table in (self.worker_rank, self.site_rank):
            table.tree.tag_configure(
                "top3", foreground=COLORS["primary"],
                font=("Microsoft YaHei UI", 9, "bold"),
            )

        # ---- Notebook 标签页 ----
        notebook = ttk.Notebook(self.parent, bootstyle=PRIMARY)
        notebook.pack(fill=BOTH, expand=True)

        detail_tab = ttk.Frame(notebook, padding=10)
        worker_tab = ttk.Frame(notebook, padding=10)
        notebook.add(detail_tab, text="  工天明细  ")
        notebook.add(worker_tab, text="  工人档案  ")

        # 工天明细：工具条（组件化，保持原左右分区与按钮顺序）
        FilterBar(
            detail_tab,
            ("搜索工人 / 工地 / 工作内容", ttk.Entry(
                detail_tab, textvariable=self.search_var, width=26
            )),
            ttk.Button(
                detail_tab, text="查询", bootstyle=INFO, command=self.refresh_logs
            ),
            ttk.Button(
                detail_tab, text="清空", bootstyle=SECONDARY, command=self.clear_search
            ),
            ttk.Button(
                detail_tab, text="切换加班标记", bootstyle="primary-outline",
                command=self.toggle_selected_overtime,
            ),
            actions=[
                ttk.Button(
                    detail_tab, text="锁定工资", bootstyle="info-outline",
                    command=lambda: self.set_selected_logs_locked(True),
                ),
                ttk.Button(
                    detail_tab, text="解除工资锁定", bootstyle="secondary-outline",
                    command=lambda: self.set_selected_logs_locked(False),
                ),
                ttk.Button(
                    detail_tab, text="删除", bootstyle=DANGER,
                    command=self.delete_selected_logs,
                ),
                ttk.Button(
                    detail_tab, text="修改", bootstyle=WARNING,
                    command=self.edit_selected_log,
                ),
            ],
        )

        # 工天明细表格（组件化，iid 取代可见 ID 列）
        self.log_table = DataTable(
            detail_tab,
            specs=(
                ("date", "日期", 90, CENTER),
                ("worker", "工人", 85, CENTER),
                ("trade", "工种", 80, CENTER),
                ("site", "施工工地", 145, CENTER),
                ("type", "工作内容", 120, CENTER),
                ("attendance", "出勤类型", 70, CENTER),
                ("days", "工天", 60, CENTER),
                ("rate", "日工资", 75, CENTER),
                ("amount", "金额", 80, CENTER),
                ("lock", "工资状态", 75, CENTER),
                ("notes", "备注", 120, CENTER),
            ),
            empty_text="本月暂无工天记录，点击右上角「新增工天」开始登记",
            stretch=("site", "type", "notes"),
        )
        self.log_tree = self.log_table.tree
        self.log_tree.configure(selectmode="extended")
        self.log_tree.tag_configure("locked", foreground=COLORS["text_muted"])
        self.log_tree.tag_configure(
            "overtime", foreground=COLORS["primary"],
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        self.log_tree.bind("<Double-1>", lambda event: self.edit_selected_log())

        # 工人档案：工具条（组件化）
        FilterBar(
            worker_tab,
            ttk.Label(
                worker_tab,
                text='工资变动请使用"调整工资"，可按生效日期批量处理历史工天。',
            ),
            actions=[
                ttk.Button(
                    worker_tab, text="新增工人", bootstyle=SUCCESS,
                    command=self.open_worker_dialog,
                ),
                ttk.Button(
                    worker_tab, text="调整工资", bootstyle=INFO,
                    command=self.open_rate_adjustment_dialog,
                ),
                ttk.Button(
                    worker_tab, text="删除工人", bootstyle=DANGER,
                    command=self.delete_selected_workers,
                ),
                ttk.Button(
                    worker_tab, text="修改工人", bootstyle=WARNING,
                    command=self.edit_selected_worker,
                ),
            ],
        )

        # 工人档案表格（组件化，iid 取代可见 ID 列）
        self.worker_table = DataTable(
            worker_tab,
            specs=(
                ("name", "姓名", 110, CENTER),
                ("trade", "工种", 110, CENTER),
                ("phone", "电话", 135, CENTER),
                ("rate", "默认日工资", 100, CENTER),
                ("status", "状态", 70, CENTER),
                ("notes", "备注", 260, CENTER),
            ),
            empty_text="暂无工人档案，点击「新增工人」建立档案",
            stretch=("name", "notes"),
        )
        self.worker_tree = self.worker_table.tree
        self.worker_tree.configure(selectmode="extended")
        self.worker_tree.bind("<Double-1>", lambda event: self.edit_selected_worker())

    @staticmethod
    def _money(value):
        return f"¥{float(value or 0):,.2f}"

    @staticmethod
    def _number(value):
        number = float(value or 0)
        return f"{number:,.2f}".rstrip("0").rstrip(".")

    @staticmethod
    def selected_id(tree):
        selected = tree.selection()
        if not selected:
            return None
        # iid 即记录 id（refresh 时以 str(id) 作为 iid）
        return int(selected[0])

    @staticmethod
    def selected_ids(tree):
        return [int(item) for item in tree.selection()]

    def refresh_months(self):
        current = datetime.now().strftime("%Y-%m")
        today = datetime.now()
        selectable = []
        for offset in range(-12, 13):
            month_index = today.month - 1 + offset
            year = today.year + month_index // 12
            selectable.append(f"{year:04d}-{month_index % 12 + 1:02d}")
        months = sorted(set(db.get_work_months() + selectable), reverse=True)
        self.month_combo["values"] = months
        if self.month_var.get() not in months:
            self.month_var.set(current)

    def refresh_all(self):
        self.refresh_kpis()
        self.refresh_ranks()
        self.refresh_logs()
        self.refresh_workers()

    def refresh_kpis(self):
        summary = db.get_work_dashboard(self.month_var.get())["summary"]
        self.kpi_vars["days"].set(self._number(summary.get("total_days")))
        self.kpi_vars["amount"].set(self._money(summary.get("total_amount")))
        self.kpi_vars["workers"].set(str(summary.get("worker_count") or 0))
        self.kpi_vars["sites"].set(str(summary.get("site_count") or 0))
        self.overtime_hint_var.set(
            f"其中加班 {summary.get('overtime_record_count') or 0} 条 / "
            f"{self._number(summary.get('overtime_days'))} 工天"
        )

    def refresh_ranks(self):
        dashboard = db.get_work_dashboard(self.month_var.get())

        def rank_mapper(row, rank, values):
            tags = ("top3",) if rank <= 3 else ()
            return None, (rank,) + values, tags

        def worker_mapper(pair):
            index, row = pair
            return rank_mapper(row, index + 1, (
                row.get("name", ""), row.get("trade", ""),
                self._number(row.get("work_days")),
                self._number(row.get("overtime_days")),
                self._money(row.get("amount")),
            ))

        def site_mapper(pair):
            index, row = pair
            return rank_mapper(row, index + 1, (
                row.get("construction_site", ""),
                row.get("worker_count", 0),
                self._number(row.get("work_days")),
                self._number(row.get("overtime_days")),
                self._money(row.get("amount")),
            ))

        self.worker_rank.refresh(list(enumerate(dashboard["by_worker"])), worker_mapper)
        self.site_rank.refresh(list(enumerate(dashboard["by_site"])), site_mapper)

    def refresh_logs(self):
        rows = db.get_work_logs(self.month_var.get(), self.search_var.get().strip())

        def mapper(row):
            locked = bool(row.get("rate_locked"))
            overtime = bool(row.get("is_overtime"))
            tags = []
            if locked:
                tags.append("locked")
            if overtime:
                tags.append("overtime")
            return str(row["id"]), (
                row.get("work_date", ""),
                row.get("worker_name", ""), row.get("trade", ""),
                row.get("construction_site", ""), row.get("work_type", ""),
                "加班" if overtime else "正常",
                self._number(row.get("work_days")),
                self._money(row.get("daily_rate")),
                self._money(row.get("amount")),
                "已锁定" if locked else "未锁定",
                row.get("notes", ""),
            ), tuple(tags)

        self.log_table.refresh(rows, mapper)

    def refresh_workers(self):
        self.worker_table.refresh(
            db.get_workers(),
            lambda worker: (str(worker["id"]), (
                worker.get("name", ""),
                worker.get("trade", ""), worker.get("phone", ""),
                self._money(worker.get("daily_rate")),
                worker.get("status", ""), worker.get("notes", ""),
            )),
        )

    def clear_search(self):
        self.search_var.set("")
        self.refresh_logs()

    def open_worker_dialog(self, worker_id=None):
        data = db.get_worker_by_id(worker_id) if worker_id else {}
        if worker_id and not data:
            messagebox.showwarning("提示", "工人不存在或已删除")
            return
        dialog = ttk.Toplevel(self.parent)
        dialog.title("修改工人" if worker_id else "新增工人")
        body, footer = build_form_dialog(
            dialog, self.parent, 620, 590, min_width=540, min_height=440
        )
        name_var = ttk.StringVar(value=data.get("name", ""))
        trade_var = ttk.StringVar(value=data.get("trade", ""))
        phone_var = ttk.StringVar(value=data.get("phone", ""))
        rate_var = ttk.StringVar(value=self._number(data.get("daily_rate", 0)))
        status_var = ttk.StringVar(value=data.get("status", "在职"))
        notes_var = ttk.StringVar(value=data.get("notes", ""))
        fields = [
            ("姓名 *", name_var), ("工种", trade_var), ("电话", phone_var),
            ("默认日工资 *", rate_var), ("备注", notes_var),
        ]
        entries = {}
        for row_index, (label, variable) in enumerate(fields):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, sticky=E, padx=(0, 12), pady=8
            )
            entry = ttk.Entry(body, textvariable=variable, width=42)
            entry.grid(row=row_index, column=1, sticky=EW, pady=8)
            entries[label] = entry
        ttk.Label(body, text="状态").grid(
            row=5, column=0, sticky=E, padx=(0, 12), pady=8
        )
        ttk.Combobox(
            body, textvariable=status_var, values=("在职", "离职"),
            state="readonly", width=39,
        ).grid(row=5, column=1, sticky=EW, pady=8)
        if worker_id:
            entries["默认日工资 *"].configure(state="disabled")
            ttk.Label(
                body, text="工资变动请回到工人档案使用“调整工资”，系统会保留历史成本。",
                bootstyle="secondary",
            ).grid(row=6, column=0, columnspan=2, sticky=W, pady=(4, 0))
        body.columnconfigure(1, weight=1)

        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("提示", "请填写姓名", parent=dialog)
                return
            try:
                rate = float(rate_var.get() or 0)
                if rate < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("提示", "默认日工资必须是非负数字", parent=dialog)
                return
            payload = {
                "name": name, "trade": trade_var.get().strip(),
                "phone": phone_var.get().strip(), "daily_rate": rate,
                "status": status_var.get(), "notes": notes_var.get().strip(),
            }
            try:
                if worker_id:
                    db.update_worker(worker_id, payload)
                else:
                    db.add_worker(payload)
            except Exception as error:
                messagebox.showwarning("无法保存", str(error), parent=dialog)
                return
            dialog.destroy()
            self.refresh_workers()

        add_form_actions(
            footer, cancel_command=dialog.destroy,
            primary_text="保存工人", primary_command=save,
        )
        entries["姓名 *"].focus_set()

    def edit_selected_worker(self):
        worker_id = self.selected_id(self.worker_tree)
        if not worker_id:
            messagebox.showwarning("提示", "请先选择一名工人")
            return
        self.open_worker_dialog(worker_id)

    def delete_selected_workers(self):
        selected = self.worker_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的工人")
            return
        ids = self.selected_ids(self.worker_tree)
        if not messagebox.askyesno("确认", f"确定删除选中的 {len(ids)} 名工人？"):
            return
        try:
            db.delete_workers(ids)
        except ValueError as error:
            messagebox.showwarning("无法删除", str(error))
            return
        self.refresh_workers()

    def open_rate_adjustment_dialog(self):
        worker_id = self.selected_id(self.worker_tree)
        if not worker_id:
            messagebox.showwarning("提示", "请先选择一名工人")
            return
        worker = db.get_worker_by_id(worker_id)
        if not worker:
            messagebox.showwarning("提示", "工人不存在或已删除")
            return
        dialog = ttk.Toplevel(self.parent)
        dialog.title(f"调整工资 · {worker['name']}")
        body, footer = build_form_dialog(
            dialog, self.parent, 790, 690, min_width=680, min_height=560
        )
        ttk.Label(body, text="调薪设置", style="CardTitle.TLabel").grid(
            row=0, column=0, columnspan=2, sticky=W, pady=(0, 10)
        )
        rate_var = ttk.StringVar(value=self._number(worker.get("daily_rate")))
        effective_var = ttk.StringVar(value=date.today().isoformat())
        end_var = ttk.StringVar(value=date.today().isoformat())
        reason_var = ttk.StringVar()
        mode_labels = {
            "仅影响今后新工天": "future_only",
            "重算生效日至今天": "through_today",
            "自定义日期和项目": "custom",
        }
        mode_var = ttk.StringVar(value="仅影响今后新工天")
        projects = project_service.list_projects(active_only=False)
        project_labels = ["全部项目"] + [
            f"{row.get('project_code') or row['id']} · {row['name']}" for row in projects
        ]
        project_var = ttk.StringVar(value="全部项目")
        project_by_label = {
            label: projects[index - 1]["id"]
            for index, label in enumerate(project_labels) if index
        }
        field_rows = [
            ("新日工资 *", rate_var), ("生效日期 *", effective_var),
        ]
        for row_index, (label, variable) in enumerate(field_rows, 1):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, sticky=E, padx=(0, 12), pady=7
            )
            if variable is effective_var:
                widget = DatePicker(
                    body,
                    textvariable=variable,
                    popup_title="选择生效日期",
                )
            else:
                widget = ttk.Entry(body, textvariable=variable, width=42)
            widget.grid(row=row_index, column=1, sticky=EW, pady=7)
        ttk.Label(body, text="影响范围 *").grid(
            row=3, column=0, sticky=E, padx=(0, 12), pady=7
        )
        mode_combo = ttk.Combobox(
            body, textvariable=mode_var, values=tuple(mode_labels), state="readonly"
        )
        mode_combo.grid(row=3, column=1, sticky=EW, pady=7)
        ttk.Label(body, text="截止日期").grid(
            row=4, column=0, sticky=E, padx=(0, 12), pady=7
        )
        end_entry = DatePicker(
            body,
            textvariable=end_var,
            popup_title="选择截止日期",
        )
        end_entry.grid(row=4, column=1, sticky=EW, pady=7)
        ttk.Label(body, text="限定项目").grid(
            row=5, column=0, sticky=E, padx=(0, 12), pady=7
        )
        project_combo = ttk.Combobox(
            body, textvariable=project_var, values=project_labels, state="readonly"
        )
        project_combo.grid(row=5, column=1, sticky=EW, pady=7)
        ttk.Label(body, text="调薪原因 *").grid(
            row=6, column=0, sticky=E, padx=(0, 12), pady=7
        )
        ttk.Entry(body, textvariable=reason_var).grid(
            row=6, column=1, sticky=EW, pady=7
        )
        ttk.Label(body, text="调整预览", style="CardTitle.TLabel").grid(
            row=7, column=0, columnspan=2, sticky=W, pady=(16, 7)
        )
        preview_box = ttk.Text(body, height=9, wrap="word")
        preview_box.grid(row=8, column=0, columnspan=2, sticky=NSEW)
        preview_box.insert("1.0", "填写调薪信息后点击“预览影响”。")
        preview_box.configure(state="disabled")
        ttk.Label(body, text="最近调薪记录", style="CardTitle.TLabel").grid(
            row=9, column=0, columnspan=2, sticky=W, pady=(16, 7)
        )
        history = ttk.Treeview(
            body, columns=("date", "rate", "scope", "count", "reason"),
            show="headings", height=5,
        )
        for column, label, width in (
            ("date", "生效日期", 95), ("rate", "新日工资", 90),
            ("scope", "范围", 130), ("count", "影响记录", 75),
            ("reason", "原因", 210),
        ):
            history.heading(column, text=label)
            history.column(column, width=width, anchor=CENTER)
        history.grid(row=10, column=0, columnspan=2, sticky=NSEW)
        for item in db.list_rate_adjustments(worker_id):
            history.insert("", END, values=(
                item.get("effective_from", ""),
                self._money((item.get("new_rate_minor") or 0) / 100),
                item.get("scope_mode", ""), item.get("affected_count", 0),
                item.get("reason", ""),
            ))
        body.columnconfigure(1, weight=1)
        body.rowconfigure(8, weight=1)

        def update_scope_state(*_):
            mode = mode_labels[mode_var.get()]
            end_entry.set_enabled(mode == "custom")
            project_combo.configure(state="readonly" if mode == "custom" else "disabled")

        mode_combo.bind("<<ComboboxSelected>>", update_scope_state)
        update_scope_state()

        def request_payload():
            mode = mode_labels[mode_var.get()]
            return {
                "worker_id": worker_id,
                "new_daily_rate": rate_var.get().strip(),
                "effective_from": effective_var.get().strip(),
                "scope_mode": mode,
                "range_end": end_var.get().strip() if mode == "custom" else None,
                "project_id": project_by_label.get(project_var.get()) if mode == "custom" else None,
                "reason": reason_var.get().strip(),
            }

        preview_cache = {"value": None}

        def show_preview():
            try:
                result = db.preview_rate_adjustment(request_payload())
            except ValueError as error:
                messagebox.showwarning("无法预览", str(error), parent=dialog)
                return None
            lines = [
                f"{worker['name']}：{self._money(result['current_rate_minor'] / 100)}"
                f" → {self._money(result['new_rate_minor'] / 100)}",
                f"{result['effective_from']} 生效 · {result['scope_label']}",
                f"影响 {result['affected_count']} 条记录、{self._number(result['total_days'])} 工天",
                f"跳过已锁定 {result['skipped_locked_count']} 条，成本变化 {self._money(result['delta_minor'] / 100)}",
            ]
            if result["project_impacts"]:
                lines.append("")
                lines.append("分项目影响：")
                for impact in result["project_impacts"]:
                    sign = "+" if impact["delta_minor"] >= 0 else ""
                    lines.append(
                        f"• {impact['project_name']}：{impact['record_count']} 条，"
                        f"{sign}{self._money(impact['delta_minor'] / 100)}"
                    )
            preview_box.configure(state="normal")
            preview_box.delete("1.0", END)
            preview_box.insert("1.0", "\n".join(lines))
            preview_box.configure(state="disabled")
            preview_cache["value"] = result
            return result

        def apply_adjustment():
            result = show_preview()
            if not result:
                return
            if not messagebox.askyesno(
                "确认调薪",
                f"确认把 {worker['name']} 的日工资调整为 "
                f"{self._money(result['new_rate_minor'] / 100)}？\n"
                f"将更新 {result['affected_count']} 条未锁定工天记录。",
                parent=dialog,
            ):
                return
            try:
                applied = db.apply_rate_adjustment(request_payload())
            except ValueError as error:
                messagebox.showwarning("无法调薪", str(error), parent=dialog)
                return
            dialog.destroy()
            self.refresh_all()
            messagebox.showinfo(
                "调薪完成", f"已更新工资版本和 {applied['affected_count']} 条工天记录"
            )

        add_form_actions(
            footer, cancel_command=dialog.destroy,
            secondary_text="预览影响", secondary_command=show_preview,
            primary_text="确认调薪", primary_command=apply_adjustment,
        )

    def _project_choices(self, include_project_id=None):
        """工天项目候选：排除已关闭项目，编辑时保留当前项目。"""
        projects = db.list_work_log_project_options(include_project_id)
        label_by_id = {
            project["id"]: f"{project['name']} · {project['project_code']}"
            for project in projects
        }
        labels = [UNASSIGNED_LABEL] + list(label_by_id.values())
        id_by_label = {label: pid for pid, label in label_by_id.items()}
        return labels, id_by_label, label_by_id

    @staticmethod
    def _site_selection(project_id):
        sites = db.list_work_log_site_options(project_id)
        names = []
        ids_by_name = {}
        duplicate_names = set()
        for site in sites:
            name = (site.get("name") or "").strip()
            if not name:
                continue
            if name in ids_by_name:
                duplicate_names.add(name)
            else:
                names.append(name)
                ids_by_name[name] = site["id"]
        for name in duplicate_names:
            ids_by_name.pop(name, None)
        return names, ids_by_name

    @classmethod
    def _configure_site_choices(
        cls, site_combo, site_var, project_id, site_ids_by_name, reset=False
    ):
        names, site_ids = cls._site_selection(project_id)
        site_ids_by_name.clear()
        site_ids_by_name.update(site_ids)
        site_combo.configure(values=names)
        if not reset or not project_id:
            return
        if len(names) == 1:
            site_var.set(names[0])
        elif site_var.get().strip() not in names:
            site_var.set("")

    @staticmethod
    def _build_work_days_selector(parent, days_var):
        """Build mutually exclusive round choices for the supported work days."""
        selector = ttk.Frame(parent)

        def keep_selected(value):
            if not days_var.get():
                days_var.set(value)

        for value, text in (("0.5", "0.5 工天"), ("1", "1 工天")):
            ttk.Checkbutton(
                selector,
                text=text,
                variable=days_var,
                onvalue=value,
                offvalue="",
                command=lambda option=value: keep_selected(option),
                bootstyle="round-toggle",
            ).pack(side=LEFT, padx=(0, 18))
        return selector

    def open_log_dialog(self, log_id=None):
        if not log_id:
            self.open_batch_log_dialog()
            return
        data = db.get_work_log_by_id(log_id)
        if not data:
            messagebox.showwarning("提示", "工天记录不存在或已作废")
            return
        workers = db.get_workers(active_only=False)
        dialog = ttk.Toplevel(self.parent)
        dialog.title("修改工天记录")
        body, footer = build_form_dialog(
            dialog, self.parent, 680, 650, min_width=580, min_height=470
        )
        worker_labels = [
            f"{worker['name']} · {worker.get('trade') or '未设工种'}" for worker in workers
        ]
        worker_by_label = {label: workers[index] for index, label in enumerate(worker_labels)}
        current_label = next(
            (label for label, worker in worker_by_label.items() if worker["id"] == data["worker_id"]),
            worker_labels[0] if worker_labels else "",
        )
        worker_var = ttk.StringVar(value=current_label)
        date_var = ttk.StringVar(value=data.get("work_date", date.today().isoformat()))
        site_var = ttk.StringVar(value=data.get("construction_site", ""))
        project_labels, id_by_label, label_by_id = self._project_choices(
            data.get("project_id")
        )
        current_project_label = label_by_id.get(data.get("project_id"))
        if not current_project_label:
            current_project_label = UNASSIGNED_LABEL
        project_var = ttk.StringVar(value=current_project_label)
        type_var = ttk.StringVar(value=data.get("work_type", ""))
        current_days = self._number(data.get("work_days", 1))
        days_var = ttk.StringVar(
            value=current_days if current_days in {"0.5", "1"} else ""
        )
        rate_var = ttk.StringVar(value=self._number(data.get("daily_rate", 0)))
        notes_var = ttk.StringVar(value=data.get("notes", ""))
        overtime_var = ttk.BooleanVar(value=bool(data.get("is_overtime")))
        amount_var = ttk.StringVar()
        widgets = []
        labels = (
            "工人 *", "日期 *", "所属项目 *", "施工地点 / 作业区域 *", "工作内容", "工天 *",
            "日工资 *", "金额", "加班标记", "备注",
        )
        for row_index, label in enumerate(labels):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, sticky=E, padx=(0, 12), pady=7
            )
        widgets.append(ttk.Combobox(body, textvariable=worker_var, values=worker_labels, state="readonly"))
        widgets.append(
            DatePicker(
                body,
                textvariable=date_var,
                popup_title="选择工天日期",
            )
        )
        widgets.append(ttk.Combobox(body, textvariable=project_var, values=project_labels, state="readonly"))
        widgets.append(ttk.Combobox(body, textvariable=site_var))
        widgets.append(ttk.Entry(body, textvariable=type_var))
        widgets.append(self._build_work_days_selector(body, days_var))
        widgets.append(ttk.Entry(body, textvariable=rate_var))
        widgets.append(ttk.Entry(body, textvariable=amount_var, state="readonly"))
        widgets.append(
            ttk.Checkbutton(
                body,
                text="标记为加班",
                variable=overtime_var,
                bootstyle="round-toggle",
            )
        )
        widgets.append(ttk.Entry(body, textvariable=notes_var))
        for row_index, widget in enumerate(widgets):
            widget.grid(row=row_index, column=1, sticky=EW, pady=7)
        body.columnconfigure(1, weight=1)

        site_ids_by_name = {}

        def refresh_site_choices(_event=None, reset=False):
            project_id = id_by_label.get(project_var.get())
            self._configure_site_choices(
                widgets[3], site_var, project_id, site_ids_by_name, reset
            )

        widgets[2].bind(
            "<<ComboboxSelected>>",
            lambda event: refresh_site_choices(event, reset=True),
        )
        refresh_site_choices()

        def update_amount(*_):
            try:
                amount_var.set(self._money(float(days_var.get()) * float(rate_var.get())))
            except ValueError:
                amount_var.set("—")

        days_var.trace_add("write", update_amount)
        rate_var.trace_add("write", update_amount)
        update_amount()

        def save():
            worker = worker_by_label.get(worker_var.get())
            if not worker or not date_var.get().strip() or not site_var.get().strip():
                messagebox.showwarning("提示", "请完整填写工人、日期和施工工地", parent=dialog)
                return
            project_label = project_var.get()
            if not project_label:
                messagebox.showwarning(
                    "提示",
                    "请选择所属项目；确实不属于任何项目时请选择“待归集”",
                    parent=dialog,
                )
                return
            if days_var.get() not in {"0.5", "1"}:
                messagebox.showwarning(
                    "提示", "请选择 0.5 工天或 1 工天", parent=dialog
                )
                return
            try:
                days = float(days_var.get())
                rate = float(rate_var.get())
                datetime.strptime(date_var.get().strip(), "%Y-%m-%d")
                if rate < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("提示", "日期或日工资格式不正确", parent=dialog)
                return
            payload = {
                "worker_id": worker["id"], "work_date": date_var.get().strip(),
                "construction_site": site_var.get().strip(),
                "work_type": type_var.get().strip(), "work_days": days,
                "daily_rate": rate, "notes": notes_var.get().strip(),
                "is_overtime": overtime_var.get(),
            }
            if project_label == UNASSIGNED_LABEL:
                if not messagebox.askyesno(
                    "确认暂不归集",
                    "这条人工记录不会进入任何项目成本，将立即出现在“数据治理中心”。\n"
                    "确定仍以待归集状态保存吗？",
                    parent=dialog,
                ):
                    widgets[2].focus_set()
                    return
                payload["allow_unassigned"] = True
            else:
                payload["project_id"] = id_by_label[project_label]
                project_site_id = site_ids_by_name.get(site_var.get().strip())
                if project_site_id:
                    payload["project_site_id"] = project_site_id
            try:
                db.update_work_log(log_id, payload)
            except ValueError as error:
                messagebox.showwarning("无法保存", str(error), parent=dialog)
                return
            except Exception as error:
                messagebox.showerror(
                    "保存失败",
                    f"记录没有保存，程序遇到异常：{error}",
                    parent=dialog,
                )
                return
            self.month_var.set(payload["work_date"][:7])
            dialog.destroy()
            self.refresh_months()
            self.refresh_all()

        add_form_actions(
            footer, cancel_command=dialog.destroy,
            primary_text="保存记录", primary_command=save,
        )

    def open_batch_log_dialog(self):
        workers = db.get_workers(active_only=True)
        if not workers:
            messagebox.showwarning("提示", "请先新增至少一名在职工人")
            self.open_worker_dialog()
            return
        dialog = ttk.Toplevel(self.parent)
        dialog.title("批量新增工天")
        body, footer = build_form_dialog(
            dialog, self.parent, 700, 690, min_width=600, min_height=480
        )
        ttk.Label(body, text="选择工人 *").grid(
            row=0, column=0, sticky=NE, padx=(0, 12), pady=7
        )
        worker_box = ttk.Frame(body)
        worker_box.grid(row=0, column=1, sticky=NSEW, pady=7)
        worker_list = Listbox(
            worker_box, selectmode=EXTENDED, height=8,
            exportselection=False, font=("Microsoft YaHei UI", 10),
        )
        worker_list.pack(fill=BOTH, expand=True)
        for worker in workers:
            worker_list.insert(
                TK_END,
                f"{worker['name']} · {worker.get('trade') or '未设工种'} · "
                f"{self._money(worker.get('daily_rate'))}/天",
            )
        tools = ttk.Frame(worker_box)
        tools.pack(fill=X, pady=(6, 0))
        ttk.Button(
            tools, text="全选", bootstyle="secondary-outline",
            command=lambda: worker_list.selection_set(0, TK_END),
        ).pack(side=LEFT)
        ttk.Button(
            tools, text="清空", bootstyle="secondary-outline",
            command=lambda: worker_list.selection_clear(0, TK_END),
        ).pack(side=LEFT, padx=6)
        date_var = ttk.StringVar(value=date.today().isoformat())
        site_var = ttk.StringVar()
        project_labels, id_by_label, _label_by_id = self._project_choices()
        project_var = ttk.StringVar()
        type_var = ttk.StringVar()
        days_var = ttk.StringVar(value="1")
        notes_var = ttk.StringVar()
        overtime_var = ttk.BooleanVar(value=False)
        fields = (
            ("日期 *", date_var, "entry"),
            ("所属项目 *", project_var, "project"),
            ("施工地点 / 作业区域 *", site_var, "site"),
            ("工作内容", type_var, "entry"),
            ("每人工天 *", days_var, "days"),
            ("备注", notes_var, "entry"),
        )
        project_combo = None
        site_combo = None
        for row_index, (label, variable, kind) in enumerate(fields, 1):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, sticky=E, padx=(0, 12), pady=7
            )
            if kind == "site":
                widget = ttk.Combobox(body, textvariable=variable)
                site_combo = widget
            elif kind == "project":
                widget = ttk.Combobox(
                    body,
                    textvariable=variable,
                    values=project_labels,
                    state="readonly",
                )
                project_combo = widget
            elif kind == "days":
                widget = self._build_work_days_selector(body, days_var)
            elif label == "日期 *":
                widget = DatePicker(
                    body,
                    textvariable=variable,
                    popup_title="选择批量工天日期",
                )
            else:
                widget = ttk.Entry(body, textvariable=variable)
            widget.grid(row=row_index, column=1, sticky=EW, pady=7)

        site_ids_by_name = {}

        def refresh_site_choices(_event=None):
            project_id = id_by_label.get(project_var.get())
            self._configure_site_choices(
                site_combo, site_var, project_id, site_ids_by_name, reset=True
            )

        if project_combo is not None:
            project_combo.bind("<<ComboboxSelected>>", refresh_site_choices)
        ttk.Label(body, text="加班标记").grid(
            row=7, column=0, sticky=E, padx=(0, 12), pady=7
        )
        ttk.Checkbutton(
            body,
            text="本批记录标记为加班",
            variable=overtime_var,
            bootstyle="round-toggle",
        ).grid(row=7, column=1, sticky=W, pady=7)
        ttk.Label(
            body,
            text="系统会按所选日期读取每名工人的生效工资；同一工人一天可拆分多个地点，"
            "但当天有效工天合计不能超过 1。选择“待归集”的人工不会计入任何项目成本。",
            bootstyle="secondary",
        ).grid(row=8, column=0, columnspan=2, sticky=W, pady=(8, 0))
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        def save():
            selected = list(worker_list.curselection())
            if not selected:
                messagebox.showwarning("提示", "请至少选择一名工人", parent=dialog)
                return
            if not date_var.get().strip() or not site_var.get().strip():
                messagebox.showwarning("提示", "请填写日期和施工工地", parent=dialog)
                return
            project_label = project_var.get()
            if not project_label:
                messagebox.showwarning(
                    "提示",
                    "请选择所属项目；确实不属于任何项目时请选择“待归集”",
                    parent=dialog,
                )
                return
            attribution = (
                {"allow_unassigned": True}
                if project_label == UNASSIGNED_LABEL
                else {"project_id": id_by_label[project_label]}
            )
            if project_label == UNASSIGNED_LABEL and not messagebox.askyesno(
                "确认暂不归集",
                f"本批 {len(selected)} 条人工记录不会进入任何项目成本，"
                "将立即出现在“数据治理中心”。\n确定仍要保存吗？",
                parent=dialog,
            ):
                if project_combo is not None:
                    project_combo.focus_set()
                return
            project_site_id = site_ids_by_name.get(site_var.get().strip())
            if project_site_id and "project_id" in attribution:
                attribution["project_site_id"] = project_site_id
            if days_var.get() not in {"0.5", "1"}:
                messagebox.showwarning(
                    "提示", "请选择 0.5 工天或 1 工天", parent=dialog
                )
                return
            try:
                work_date = datetime.strptime(
                    date_var.get().strip(), "%Y-%m-%d"
                ).date().isoformat()
                days = float(days_var.get())
            except ValueError:
                messagebox.showwarning("提示", "日期格式不正确", parent=dialog)
                return
            selected_workers = [workers[index] for index in selected]
            try:
                rates = db.get_effective_worker_rates(
                    [worker["id"] for worker in selected_workers], work_date
                )
                entries = [
                    {
                        "worker_id": worker["id"], "work_date": work_date,
                        "construction_site": site_var.get().strip(),
                        "work_type": type_var.get().strip(), "work_days": days,
                        "daily_rate": rates[worker["id"]],
                        "notes": notes_var.get().strip(),
                        "is_overtime": overtime_var.get(),
                        **attribution,
                    }
                    for worker in selected_workers
                ]
                count = db.add_work_logs_batch(entries)
            except ValueError as error:
                messagebox.showwarning("无法保存", str(error), parent=dialog)
                return
            except Exception as error:
                messagebox.showerror(
                    "保存失败",
                    f"本批记录没有保存，程序遇到异常：{error}",
                    parent=dialog,
                )
                return
            self.month_var.set(work_date[:7])
            dialog.destroy()
            self.refresh_months()
            self.refresh_all()
            messagebox.showinfo("保存成功", f"已为 {count} 名工人生成工天记录")

        add_form_actions(
            footer, cancel_command=dialog.destroy,
            primary_text="批量保存", primary_command=save,
        )

    def edit_selected_log(self):
        log_id = self.selected_id(self.log_tree)
        if not log_id:
            messagebox.showwarning("提示", "请先选择一条工天记录")
            return
        self.open_log_dialog(log_id)

    def toggle_selected_overtime(self):
        selected = self.log_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择工天记录")
            return
        log_ids = self.selected_ids(self.log_tree)
        rows = [db.get_work_log_by_id(log_id) for log_id in log_ids]
        mark_as_overtime = not all(
            row and row.get("is_overtime") for row in rows
        )
        changed = db.set_work_logs_overtime(log_ids, mark_as_overtime)
        self.refresh_all()
        action = "标记为加班" if mark_as_overtime else "取消加班标记"
        messagebox.showinfo("操作完成", f"已为 {changed} 条工天记录{action}")

    def set_selected_logs_locked(self, locked):
        selected = self.log_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择工天记录")
            return
        ids = self.selected_ids(self.log_tree)
        action = "锁定" if locked else "解除锁定"
        reason = simpledialog.askstring(
            f"{action}工资", f"请输入{action}原因：", parent=self.parent
        )
        if reason is None:
            return
        try:
            changed = db.set_work_logs_rate_locked(ids, locked, reason)
        except ValueError as error:
            messagebox.showwarning("操作失败", str(error))
            return
        self.refresh_logs()
        messagebox.showinfo("操作完成", f"已{action} {changed} 条工天记录")

    def delete_selected_logs(self):
        selected = self.log_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的工天记录")
            return
        ids = self.selected_ids(self.log_tree)
        if not messagebox.askyesno("确认", f"确定删除选中的 {len(ids)} 条工天记录？"):
            return
        try:
            db.delete_work_logs(ids)
        except ValueError as error:
            messagebox.showwarning("无法删除", str(error))
            return
        self.refresh_all()

    def export_month(self):
        month = self.month_var.get()
        rows = db.get_work_logs(month)
        if not rows:
            messagebox.showinfo("提示", f"{month} 暂无工天记录")
            return
        dashboard = db.get_work_dashboard(month)
        path = filedialog.asksaveasfilename(
            title="导出月度工天表", defaultextension=".xlsx",
            initialfile=f"{month}_工天统计.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        workbook = Workbook()
        detail_sheet = workbook.active
        detail_sheet.title = "工天明细"
        detail_sheet.append([
            "日期", "工人", "工种", "施工工地", "工作内容",
            "出勤类型", "工天", "日工资", "金额", "工资状态", "备注",
        ])
        for row in rows:
            detail_sheet.append([
                row.get("work_date", ""), row.get("worker_name", ""),
                row.get("trade", ""), row.get("construction_site", ""),
                row.get("work_type", ""),
                "加班" if row.get("is_overtime") else "正常",
                float(row.get("work_days") or 0),
                float(row.get("daily_rate") or 0), float(row.get("amount") or 0),
                "已锁定" if row.get("rate_locked") else "未锁定",
                row.get("notes", ""),
            ])
        worker_sheet = workbook.create_sheet("按工人汇总")
        worker_sheet.append([
            "工人", "工种", "工天", "加班工天", "人工费", "施工地数量"
        ])
        for row in dashboard["by_worker"]:
            worker_sheet.append([
                row.get("name", ""), row.get("trade", ""),
                float(row.get("work_days") or 0),
                float(row.get("overtime_days") or 0),
                float(row.get("amount") or 0),
                int(row.get("site_count") or 0),
            ])
        site_sheet = workbook.create_sheet("按工地汇总")
        site_sheet.append(["施工工地", "人数", "工天", "加班工天", "人工费"])
        for row in dashboard["by_site"]:
            site_sheet.append([
                row.get("construction_site", ""), int(row.get("worker_count") or 0),
                float(row.get("work_days") or 0),
                float(row.get("overtime_days") or 0),
                float(row.get("amount") or 0),
            ])
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 3, 34)
                sheet.column_dimensions[column[0].column_letter].width = width
        try:
            workbook.save(path)
        except Exception as error:
            messagebox.showerror("导出失败", str(error))
            return
        messagebox.showinfo("导出完成", f"已保存到：\n{path}")
